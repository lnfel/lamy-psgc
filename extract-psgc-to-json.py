from openpyxl import load_workbook
import json
import io
import sys
from pathlib import Path

if len(sys.argv) > 1:
    filePath = Path(sys.argv[1])
else:
    # Default file used for testing
    filePath = Path('./PSGC 4Q 2022 Publication Datafile.xlsx')

if not filePath.is_file():
    print('Please provide path to PSGC publication file (Excel file) or make sure the path is correct.')
    print('Example: python3 ./extract-psgc-to-json.py "PSGC 4Q 2022 Publication Datafile.xlsx"')
    exit()

sourceFile = filePath

# Load entire workbook
wb = load_workbook(sourceFile, data_only=True)

# List all the sheets in the file
print('Found the following sheets:')

for sheetname in wb.sheetnames:
    print(sheetname)

# Load one worksheet
ws = wb['PSGC']

allRows = list(ws.rows)
print(f"\nFound {len(allRows)} rows of data.")

# Extract rows base on Geographic Level ('Reg', 'Prov', 'Mun', 'City', 'Bgy')
def filterGeographicLevel(allRows, geographicLevel):
    """Filter rows based on Geographic Level and get all rows"""

    belongsToRegion = ['Prov', 'Mun', 'City', 'Bgy']
    belongsToProvince = ['Mun', 'City', 'Bgy']
    # belongs to municipality or city
    belongsToMunicipality = ['Bgy']

    data = []
    for row in allRows:
        if row[3].value == geographicLevel:
            psgcTenDigit = str(row[0].value)
            correspondenceCode = str(row[2].value)
            area = {
                "name": row[1].value,
                "psgcTenDigit": psgcTenDigit,
                "correspondenceCode": correspondenceCode,
                "geographicLevel": row[3].value
            }

            # note that municipality and city are using the same slot in the 10 digit psgc code
            match geographicLevel:
                case "Reg":
                    area['code'] = psgcTenDigit[:2]
                case "Prov":
                    area['code'] = psgcTenDigit[2:5]
                case "Mun":
                    area['code'] = psgcTenDigit[5:7]
                case "City":
                    area['code'] = psgcTenDigit[5:7]
                case "Bgy":
                    area['code'] = psgcTenDigit[7:10]

            if geographicLevel in belongsToRegion:
                # region_code is first 2 digit of psgcTenDigit
                area['region_code'] = psgcTenDigit[:2]
            if geographicLevel in belongsToProvince:
                # province_code is three digit starting from 3rd digit of psgcTenDigit
                area['province_code'] = psgcTenDigit[2:5]
            if geographicLevel in belongsToMunicipality:
                # municipality can also be a city
                area['municipality_code'] = psgcTenDigit[5:7]

            data.append(area)

    return data

# Iteration helpers
geographicLevels = ['region', 'province', 'municipality', 'district', 'city', 'barangay']
outputFolder = 'database/'
geographicLevelMap = {
    "region": {
        "filterValue": "Reg",
        "filename": "regions.json"
    },
    "province": {
        "filterValue": "Prov",
        "filename": "provinces.json"
    },
    "municipality": {
        "filterValue": "Mun",
        "filename": "municipalities.json"
    },
    "district": {
        "filterValue": "Dist",
        "filename": "districts.json"
    },
    "city": {
        "filterValue": "City",
        "filename": "cities.json"
    },
    "barangay": {
        "filterValue": "Bgy",
        "filename": "barangays.json"
    },
}

# Writing to json file
for geographicLevel in geographicLevels:
    filepath = outputFolder + geographicLevelMap[geographicLevel]['filename']
    with io.open(filepath, 'w', encoding='utf8') as file:
        data = json.dumps(
            filterGeographicLevel(allRows, geographicLevelMap[geographicLevel]['filterValue']),
            indent=4,
            sort_keys=True,
            separators=(',', ': '),
            ensure_ascii=False)
        file.write(str(data))
        print(f"Saving to {filepath}")